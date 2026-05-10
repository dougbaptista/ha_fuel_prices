import aiohttp
import zipfile
import unicodedata
import os
import tempfile
from datetime import timedelta
from bs4 import BeautifulSoup
import re
import logging

from openpyxl import load_workbook

from homeassistant.components.sensor import SensorEntity
from homeassistant.config_entries import ConfigEntry
from homeassistant.core import HomeAssistant
from homeassistant.helpers.update_coordinator import (
    DataUpdateCoordinator,
    UpdateFailed,
)

from .const import DOMAIN

BASE_URL = (
    "https://www.gov.br/anp/pt-br/assuntos/precos-e-defesa-da-concorrencia/"
    "precos/levantamento-de-precos-de-combustiveis-ultimas-semanas-pesquisadas"
)
logger = logging.getLogger(__name__)

# Intervalo de atualização: dados da ANP são semanais, 6 horas é mais que suficiente
UPDATE_INTERVAL = timedelta(hours=6)


def normalize_text(text):
    """Remove acentos, converte para ASCII, remove espaços e coloca em uppercase."""
    if not isinstance(text, str):
        text = str(text)
    normalized = unicodedata.normalize("NFD", text).encode("ascii", "ignore").decode("utf-8")
    return normalized.upper().strip()


async def async_setup_entry(hass: HomeAssistant, entry: ConfigEntry, async_add_entities):
    """Configuração inicial do sensor no Home Assistant."""

    coordinator = FuelPriceCoordinator(hass, entry.data)
    await coordinator.async_config_entry_first_refresh()

    fuel_types = [
        "ETANOL HIDRATADO",
        "GASOLINA COMUM",
        "GASOLINA ADITIVADA",
        "GLP",
        "GNV",
        "OLEO DIESEL",
        "OLEO DIESEL S10",
    ]
    sensors = []
    price_type_names = {"min": "Mínimo", "med": "Médio", "max": "Máximo"}

    for fuel in fuel_types:
        for pt in ["min", "med", "max"]:
            city = entry.data.get("city", "TUBARAO").title()
            sensor_name = f"Preço {fuel.title()} ({city}) - {price_type_names[pt]}"
            sensors.append(FuelPriceSensor(coordinator, entry.data, fuel, pt, sensor_name))

    async_add_entities(sensors)


class FuelPriceCoordinator(DataUpdateCoordinator):
    """Coordenador que busca os dados da ANP uma única vez e compartilha com todos os sensores."""

    def __init__(self, hass: HomeAssistant, config: dict):
        super().__init__(
            hass,
            logger,
            name=DOMAIN,
            update_interval=UPDATE_INTERVAL,
        )
        self._config = config

    async def _async_update_data(self) -> dict:
        """Busca os dados da ANP. Retorna dicionário de preços por combustível."""
        try:
            xls_url = await fetch_latest_xls_url()
            if not xls_url:
                raise UpdateFailed("Não foi possível encontrar a URL XLS mais recente.")

            prices = await download_and_extract_sc_prices(xls_url, self._config)
            if not prices:
                raise UpdateFailed(
                    f"Nenhum dado retornado para "
                    f"{self._config.get('state', '')} / {self._config.get('city', '')}"
                )
            return prices
        except UpdateFailed:
            raise
        except Exception as e:
            raise UpdateFailed(f"Erro ao buscar dados da ANP: {e}") from e


class FuelPriceSensor(SensorEntity):
    """Sensor de preço de combustível para um tipo específico de preço."""

    def __init__(self, coordinator: FuelPriceCoordinator, config, fuel_type, price_type, sensor_name):
        self._coordinator = coordinator
        self._fuel_type = fuel_type
        self._price_type = price_type
        self._config = config
        self._last_valid_state = None
        self._attr_name = sensor_name
        self._attr_unique_id = (
            f"{DOMAIN}_{fuel_type.lower().replace(' ', '_')}_"
            f"{price_type}_{config.get('city', '').lower().replace(' ', '_')}"
        )
        self._attr_unit_of_measurement = "BRL/L" if fuel_type != "GLP" else "BRL/kg"
        self._attr_extra_state_attributes = {}

        # Ícones por tipo de combustível
        icon_map = {
            "ETANOL HIDRATADO": "mdi:gas-station",
            "GASOLINA COMUM": "mdi:gas-station",
            "GASOLINA ADITIVADA": "mdi:gas-station",
            "GLP": "mdi:gas-cylinder",
            "GNV": "mdi:gas-cylinder",
            "OLEO DIESEL": "mdi:oil",
            "OLEO DIESEL S10": "mdi:oil",
        }
        self._attr_icon = icon_map.get(fuel_type, "mdi:cash")

    @property
    def should_poll(self) -> bool:
        """Desativa polling individual; o coordinator controla a atualização."""
        return False

    @property
    def available(self) -> bool:
        """Sensor disponível se o coordinator tiver dados ou se houver último valor válido."""
        return self._coordinator.last_update_success or self._last_valid_state is not None

    @property
    def native_value(self):
        return self._last_valid_state

    async def async_added_to_hass(self):
        """Registra o sensor para receber atualizações do coordinator."""
        self.async_on_remove(
            self._coordinator.async_add_listener(self._handle_coordinator_update)
        )
        # Processa dados iniciais se já disponíveis
        self._handle_coordinator_update()

    def _handle_coordinator_update(self):
        """Processa novos dados do coordinator, mantendo último valor válido."""
        if self._coordinator.data is None:
            # Coordinator falhou, mantém último valor válido
            self.async_write_ha_state()
            return

        price_info = self._coordinator.data.get(self._fuel_type)
        if price_info:
            new_value = price_info.get(self._price_type)
            if new_value is not None and isinstance(new_value, (int, float)):
                self._last_valid_state = new_value
                self._attr_extra_state_attributes = {
                    "min": price_info.get("min"),
                    "médio": price_info.get("med"),
                    "máximo": price_info.get("max"),
                }
            else:
                logger.warning(
                    f"Valor inválido para {self._fuel_type} ({self._price_type}): "
                    f"{new_value}. Mantendo último valor válido."
                )
        else:
            logger.warning(
                f"Produto {self._fuel_type} não encontrado nos dados. "
                f"Mantendo último valor válido."
            )

        self.async_write_ha_state()


async def fetch_latest_xls_url():
    """Busca a URL do último XLS cujo link contenha o texto esperado."""
    async with aiohttp.ClientSession() as session:
        async with session.get(BASE_URL) as response:
            if response.status != 200:
                raise ValueError(f"Falha ao acessar a página da ANP: {response.status}")
            html = await response.text()

    soup = BeautifulSoup(html, "html.parser")
    links = soup.find_all("a", text=re.compile(
        r"Preços médios semanais: Brasil, regiões, estados e municípios"
    ))
    if not links:
        return None
    latest_link = links[0]["href"]
    if latest_link.startswith("/"):
        latest_link = "https://www.gov.br" + latest_link
    return latest_link


async def download_and_extract_sc_prices(xls_url, config):
    """
    Baixa o arquivo XLSX e retorna um dicionário com os preços para cada combustível,
    para o município e estado configurados, a partir da aba "MUNICIPIOS".
    Usa openpyxl diretamente (sem pandas) para menor consumo de memória.
    """
    temp_dir = tempfile.gettempdir()
    temp_path = os.path.join(temp_dir, "fuel_prices_sc.xlsx")
    logger.debug(f"Arquivo XLSX será salvo em: {temp_path}")

    async with aiohttp.ClientSession() as session:
        async with session.get(xls_url) as response:
            if response.status != 200:
                raise ValueError(f"Falha ao baixar o arquivo XLS: {response.status}")
            content = await response.read()
            with open(temp_path, "wb") as f:
                f.write(content)

    if not zipfile.is_zipfile(temp_path):
        logger.error("O arquivo XLSX baixado está corrompido ou incompleto (ZIP inválido).")
        return {}

    try:
        wb = load_workbook(temp_path, read_only=True, data_only=True)
        if "MUNICIPIOS" not in wb.sheetnames:
            raise ValueError("Aba 'MUNICIPIOS' não encontrada no arquivo.")
        ws = wb["MUNICIPIOS"]

        # Linha 10 (índice 10, pois skiprows=9) é o cabeçalho
        header_row = 10
        headers = []
        for cell in ws[header_row]:
            headers.append(normalize_text(str(cell.value)) if cell.value else "")

        logger.debug("Cabeçalhos do XLSX: " + ", ".join(headers))

        # Mapeia índices das colunas necessárias
        col_map = {}
        required_cols = ["ESTADO", "MUNICIPIO", "PRODUTO", "PRECO MEDIO REVENDA",
                         "PRECO MINIMO REVENDA", "PRECO MAXIMO REVENDA"]
        for col_name in required_cols:
            if col_name in headers:
                col_map[col_name] = headers.index(col_name)
            else:
                raise ValueError(f"Coluna esperada '{col_name}' não foi encontrada.")

        state_filter = normalize_text(config.get("state", "SANTA CATARINA"))
        city_filter = normalize_text(config.get("city", "TUBARAO"))

        prices = {}
        found_any = False

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            estado = normalize_text(str(row[col_map["ESTADO"]])) if row[col_map["ESTADO"]] else ""
            municipio = normalize_text(str(row[col_map["MUNICIPIO"]])) if row[col_map["MUNICIPIO"]] else ""

            if estado != state_filter or municipio != city_filter:
                continue

            found_any = True
            product = str(row[col_map["PRODUTO"]]).strip() if row[col_map["PRODUTO"]] else ""

            try:
                raw_med = row[col_map["PRECO MEDIO REVENDA"]]
                raw_min = row[col_map["PRECO MINIMO REVENDA"]]
                raw_max = row[col_map["PRECO MAXIMO REVENDA"]]
                price_med = float(str(raw_med).replace(",", ".")) if raw_med is not None else None
                price_min = float(str(raw_min).replace(",", ".")) if raw_min is not None else None
                price_max = float(str(raw_max).replace(",", ".")) if raw_max is not None else None
            except Exception as e:
                logger.error(f"Erro ao converter os preços para o produto {product}: {e}")
                price_med = price_min = price_max = None

            logger.debug(f"Produto: {product}, min: {price_min}, med: {price_med}, max: {price_max}")
            prices[product] = {"min": price_min, "med": price_med, "max": price_max}

        wb.close()

        if not found_any:
            logger.error(
                f"Nenhum registro encontrado para {state_filter} / {city_filter} na aba MUNICIPIOS."
            )
            return {}

        return prices
    except Exception as e:
        raise ValueError(f"Erro ao processar a aba 'MUNICIPIOS': {e}")
